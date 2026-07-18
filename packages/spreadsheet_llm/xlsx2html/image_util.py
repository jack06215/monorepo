"""Extraction of images embedded in a worksheet.

Ported from the retired xlsx2md package, with one fix: the original mapped
anchor columns via string.ascii_uppercase[col], which breaks past column Z;
this version uses openpyxl's get_column_letter.
"""

import base64
import io
from typing import Callable

import openpyxl.worksheet.worksheet
from openpyxl.utils import get_column_letter
from PIL import Image


def get_pil_image_as_base64(image: Image.Image) -> str:
    """Return a Pillow image re-encoded as a base64 PNG string."""
    with io.BytesIO() as buffer:
        image.save(buffer, format="PNG")
        return base64.b64encode(buffer.getvalue()).decode()


class SheetImageLoader:
    """Indexes all images in a sheet by their anchor cell coordinate."""

    def __init__(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        self._images: dict[str, Callable[[], bytes]] = {}
        for image in sheet._images:  # type: ignore[attr-defined]
            anchor = image.anchor._from
            coordinate = f"{get_column_letter(anchor.col + 1)}{anchor.row + 1}"
            self._images[coordinate] = image._data

    def coordinates(self) -> list[str]:
        """All cell coordinates that anchor an image."""
        return list(self._images)

    def has_image(self, coordinate: str) -> bool:
        return coordinate in self._images

    def get_image_base64(self, coordinate: str) -> str:
        """Retrieve the image anchored at `coordinate` as base64 PNG."""
        if coordinate not in self._images:
            raise ValueError(f"Cell {coordinate} doesn't contain an image")
        with io.BytesIO(self._images[coordinate]()) as raw:
            return get_pil_image_as_base64(Image.open(raw))
