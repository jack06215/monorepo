"""Converts an Excel file to token-compact, style-preserving HTML.

Reads everything through openpyxl's native API (no pandas round-trip), so
cell metadata that markdown conversion destroyed — font color/weight, fill
color, borders, merged ranges, number formats, hyperlinks, comments,
hidden rows/columns — survives into the output. The HTML is optimized for
LLM consumption: repeated styles are deduplicated into CSS classes, only
deviations from a plain default cell are emitted, and embedded images
become "[image]" placeholders (base64 kept on the model, not in the HTML).

Known limitations (accepted):
  - Conditional-formatting *results* are not evaluated (openpyxl only
    exposes the rules, not which cells currently match).
  - data_only=True renders cached formula values; a workbook saved without
    a calculation cache shows formula cells as empty.
  - Charts and shapes are not captured (only embedded images).
"""

import datetime
import html
from typing import Optional, cast

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import (column_index_from_string,
                                 coordinate_from_string)
from openpyxl.worksheet.worksheet import Worksheet

from packages.spreadsheet_llm.xlsx2html import types
from packages.spreadsheet_llm.xlsx2html.image_util import SheetImageLoader
from packages.spreadsheet_llm.xlsx2html.style_util import (StyleRegistry,
                                                         cell_css,
                                                         extract_theme_palette)


def _has_time_part(number_format: str) -> bool:
    fmt = number_format.lower()
    return "h" in fmt or "s" in fmt


def _render_value(value: object, number_format: str) -> str:
    """Render a cell value as display text, using the number format only
    where it changes meaning (dates, percentages) — values are otherwise
    kept verbatim."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime.datetime):
        if value.time() == datetime.time(0, 0) and not _has_time_part(number_format):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (int, float)) and "%" in number_format:
        percent = round(float(value) * 100, 4)
        return f"{int(percent) if percent.is_integer() else percent}%"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


class Xlsx2Html:
    """Converts an Excel file to one HTML fragment per worksheet."""

    def __init__(self, filename: str) -> None:
        self._workbook = openpyxl.load_workbook(filename, data_only=True)
        self._palette = extract_theme_palette(self._workbook)

    def parse(self) -> list[types.ParsedWorksheet]:
        """Convert every worksheet to a ParsedWorksheet."""
        return [self._parse_sheet(sheet) for sheet in self._workbook.worksheets]

    def _parse_sheet(self, sheet: Worksheet) -> types.ParsedWorksheet:
        image_loader = SheetImageLoader(sheet)
        registry = StyleRegistry()

        # Merged ranges: the anchor gets colspan/rowspan, covered cells are
        # not emitted at all.
        spans: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for merged_range in sheet.merged_cells.ranges:
            anchor = (merged_range.min_row, merged_range.min_col)
            spans[anchor] = (
                merged_range.max_row - merged_range.min_row + 1,
                merged_range.max_col - merged_range.min_col + 1,
            )
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if (row, col) != anchor:
                        covered.add((row, col))

        # Pass 1: extract per-cell CSS and find the real used bounds so
        # trailing rows/columns that are entirely empty and unstyled are
        # trimmed (max_row/max_column routinely overshoot).
        css_by_cell: dict[tuple[int, int], dict[str, str]] = {}
        n_rows = n_cols = 0
        for row in range(1, (sheet.max_row or 0) + 1):
            for col in range(1, (sheet.max_column or 0) + 1):
                # sheet.cell returns Cell | MergedCell; MergedCell quacks
                # like Cell for everything read here.
                cell = cast(Cell, sheet.cell(row=row, column=col))
                css = cell_css(cell, self._palette)
                if css:
                    css_by_cell[(row, col)] = css
                significant = (
                    cell.value is not None
                    or bool(css)
                    or image_loader.has_image(cell.coordinate)
                )
                if significant:
                    n_rows = max(n_rows, row)
                    n_cols = max(n_cols, col)

        # Image anchors can sit outside the cell-data bounds (max_row and
        # max_column only track cells) — extend the grid to include them.
        for coordinate in image_loader.coordinates():
            col_str, row_idx = coordinate_from_string(coordinate)
            n_rows = max(n_rows, row_idx)
            n_cols = max(n_cols, column_index_from_string(col_str))

        # Pass 2: emit rows.
        images: list[str] = []
        body: list[str] = []
        for row in range(1, n_rows + 1):
            dimension = sheet.row_dimensions.get(row)
            hidden = dimension is not None and bool(dimension.hidden)
            body.append('<tr class="hidden">' if hidden else "<tr>")
            for col in range(1, n_cols + 1):
                if (row, col) in covered:
                    continue
                cell = cast(Cell, sheet.cell(row=row, column=col))
                body.append(
                    self._render_cell(
                        cell,
                        registry.class_for(css_by_cell.get((row, col), {})),
                        spans.get((row, col)),
                        image_loader,
                        images,
                    )
                )
            body.append("</tr>")

        table_attrs = [f'data-sheet="{html.escape(str(sheet.title), quote=True)}"']
        hidden_cols = self._hidden_columns(sheet, n_cols)
        if hidden_cols:
            table_attrs.append(f'data-hidden-cols="{",".join(hidden_cols)}"')
        table = f'<table {" ".join(table_attrs)}>{"".join(body)}</table>'

        style_block = registry.style_block()
        fragment = f"{style_block}\n{table}" if style_block else table
        return types.ParsedWorksheet(
            worksheet_name=str(sheet.title),
            html=fragment,
            base64_encoded_images=images,
        )

    def _render_cell(
        self,
        cell: Cell,
        css_class: Optional[str],
        span: Optional[tuple[int, int]],
        image_loader: SheetImageLoader,
        images: list[str],
    ) -> str:
        attrs = []
        if css_class:
            attrs.append(f'class="{css_class}"')
        if span:
            rowspan, colspan = span
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
        number_format = cell.number_format
        if number_format and number_format != "General":
            attrs.append(f'data-nf="{html.escape(number_format, quote=True)}"')
        # comment/hyperlink are missing from openpyxl-stubs' Cell.
        comment = cell.comment  # type: ignore[attr-defined]
        if comment is not None and comment.text:
            attrs.append(f'title="{html.escape(comment.text, quote=True)}"')

        content = html.escape(_render_value(cell.value, number_format))
        hyperlink = cell.hyperlink  # type: ignore[attr-defined]
        if hyperlink is not None:
            target = hyperlink.target or f"#{hyperlink.location}"
            content = f'<a href="{html.escape(target, quote=True)}">{content}</a>'
        if image_loader.has_image(cell.coordinate):
            images.append(image_loader.get_image_base64(cell.coordinate))
            content = f"{content} [image]" if content else "[image]"

        attr_text = f' {" ".join(attrs)}' if attrs else ""
        return f"<td{attr_text}>{content}</td>"

    @staticmethod
    def _hidden_columns(sheet: Worksheet, n_cols: int) -> list[str]:
        hidden = []
        for letter, dimension in sheet.column_dimensions.items():
            if dimension.hidden and dimension.min is not None:
                if dimension.min <= n_cols:
                    hidden.append(letter)
        return sorted(hidden)
