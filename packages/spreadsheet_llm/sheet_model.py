"""
sheet_model.py

Loads a real .xlsx worksheet (via openpyxl) into a simple in-memory grid of
Cell objects that the rest of SheetCompressor operates on.

This is the "Section 3.1 vanilla" data model: S = {Cell_i,j} for i in m, j in n.
"""

import colorsys
from dataclasses import dataclass
from typing import List, Optional
from xml.etree import ElementTree

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter


@dataclass
class Cell:
    row: int  # 1-indexed row
    col: int  # 1-indexed column
    value: object  # raw python value (str/int/float/date/None)
    number_format: str  # Excel's NFS, e.g. "yyyy-mm-dd", "#,##0.00", "General"
    is_merged_anchor: bool = (
        False  # True if this is the top-left cell of a merged range
    )
    merged_range: Optional[str] = None
    # Visual formatting, resolved to concrete values at load time. Colors are
    # "#RRGGBB" with visually-default ones (black text, white/no fill) already
    # normalized to None, so a fully-default cell carries all falsy fields.
    fill_color: Optional[str] = None
    font_color: Optional[str] = None
    bold: bool = False
    italic: bool = False
    underline: bool = False
    strike: bool = False
    font_size: Optional[float] = None

    @property
    def address(self) -> str:
        return f"{get_column_letter(self.col)}{self.row}"

    @property
    def is_empty(self) -> bool:
        return self.value is None or (
            isinstance(self.value, str) and self.value.strip() == ""
        )

    def style_signature(self, base_font_size: Optional[float] = None) -> Optional[str]:
        """
        Compact human/LLM-readable description of this cell's non-default
        formatting, e.g. "fill:#4472C4 font:#FFFFFF bold 14pt", or None if
        the cell looks plain. Font size is only mentioned when it deviates
        from `base_font_size` (pass the sheet's prevailing size so only
        outliers -- titles, headers -- get annotated).
        """
        parts = []
        if self.fill_color:
            parts.append(f"fill:{self.fill_color}")
        if self.font_color:
            parts.append(f"font:{self.font_color}")
        if self.bold:
            parts.append("bold")
        if self.italic:
            parts.append("italic")
        if self.underline:
            parts.append("underline")
        if self.strike:
            parts.append("strikethrough")
        if (
            self.font_size
            and base_font_size is not None
            and self.font_size != base_font_size
        ):
            parts.append(f"{self.font_size:g}pt")
        return " ".join(parts) if parts else None


# ---------------------------------------------------------------------------
# Color resolution: openpyxl Color objects come in three flavors (rgb,
# indexed, theme+tint). Normalize all of them to "#RRGGBB" so downstream
# consumers (and the LLM) see a concrete color instead of "theme 4".
# ---------------------------------------------------------------------------

# Order in which cell-style theme indices map onto the theme XML's
# <a:clrScheme> children. Note the lt1/dk1 and lt2/dk2 swaps vs document
# order -- this is Excel's actual index mapping (0 = background 1, ...).
_THEME_COLOR_ORDER = [
    "lt1", "dk1", "lt2", "dk2",
    "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
]
_DRAWML_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def _theme_palette(wb) -> List[str]:
    """Extract the workbook theme's color scheme as 'RRGGBB' strings."""
    theme_xml = getattr(wb, "loaded_theme", None)
    if not theme_xml:
        return []
    try:
        root = ElementTree.fromstring(theme_xml)
        scheme = root.find(f"{_DRAWML_NS}themeElements/{_DRAWML_NS}clrScheme")
        if scheme is None:
            return []
        palette = []
        for name in _THEME_COLOR_ORDER:
            el = scheme.find(f"{_DRAWML_NS}{name}")
            if el is None:
                return []
            child = list(el)[0]  # <a:srgbClr val=".."/> or <a:sysClr .. lastClr="..">
            palette.append(child.attrib.get("lastClr") or child.attrib["val"])
        return palette
    except (ElementTree.ParseError, AttributeError, IndexError, KeyError):
        return []


def _apply_tint(rgb_hex: str, tint: float) -> str:
    """Excel tint: scale HLS luminance toward white (>0) or black (<0)."""
    if not tint:
        return rgb_hex
    r, g, b = (int(rgb_hex[i : i + 2], 16) / 255.0 for i in (0, 2, 4))
    h, lum, s = colorsys.rgb_to_hls(r, g, b)
    lum = lum * (1.0 + tint) if tint < 0 else lum * (1.0 - tint) + tint
    r, g, b = colorsys.hls_to_rgb(h, min(max(lum, 0.0), 1.0), s)
    return f"{round(r * 255):02X}{round(g * 255):02X}{round(b * 255):02X}"


def _resolve_color(color, theme_palette: List[str]) -> Optional[str]:
    """openpyxl Color -> '#RRGGBB', or None when unset/system-default."""
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        v = color.rgb
        if not isinstance(v, str) or len(v) < 6:
            return None
        if len(v) == 8 and v[:2] == "00":  # fully transparent
            return None
        return f"#{v[-6:].upper()}"
    if ctype == "indexed":
        idx = color.indexed
        # 64/65 are the system foreground/background pseudo-indices.
        if idx is None or idx in (64, 65) or idx >= len(COLOR_INDEX):
            return None
        return f"#{COLOR_INDEX[idx][-6:].upper()}"
    if ctype == "theme":
        idx = color.theme
        if not theme_palette or not isinstance(idx, int) or idx >= len(theme_palette):
            return None
        return f"#{_apply_tint(theme_palette[idx], color.tint or 0.0).upper()}"
    return None


class Sheet:
    """A dense m x n grid of Cell objects, 1-indexed to match Excel addressing."""

    def __init__(
        self, cells: List[List[Cell]], n_rows: int, n_cols: int, name: str = "Sheet1"
    ):
        self.cells = cells  # cells[row-1][col-1] -> Cell
        self.n_rows = n_rows
        self.n_cols = n_cols
        self.name = name

    def get(self, row: int, col: int) -> Cell:
        return self.cells[row - 1][col - 1]

    def __repr__(self):
        return f"Sheet(name={self.name!r}, rows={self.n_rows}, cols={self.n_cols})"


def load_xlsx(path: str, sheet_name: Optional[str] = None) -> Sheet:
    """
    Load a worksheet from an xlsx file into a Sheet of Cells.

    - Uses openpyxl with data_only=True so formulas resolve to their last
      calculated values (falls back to formula string if no cached value).
    - Fills in every cell in the used range, including empty ones, so the
      grid is dense (required for structural-anchor / aggregation steps).
    - Propagates merged-cell values to every cell in the merged range, since
      openpyxl only stores the value on the top-left cell of a merge; this
      matches how a person visually reads a merged header.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    theme = _theme_palette(wb)

    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0

    # Pre-fetch number formats + values + formatting in one pass
    grid: List[List[Cell]] = []
    for r in range(1, n_rows + 1):
        row_cells = []
        for c in range(1, n_cols + 1):
            src = ws.cell(row=r, column=c)
            font = src.font
            fill = src.fill
            fill_color = None
            if fill is not None and fill.patternType is not None:
                fill_color = _resolve_color(fill.fgColor, theme)
                if fill_color == "#FFFFFF":  # solid white = visually no fill
                    fill_color = None
            font_color = _resolve_color(font.color, theme) if font else None
            if font_color == "#000000":  # explicit black = default text color
                font_color = None
            row_cells.append(
                Cell(
                    row=r,
                    col=c,
                    value=src.value,
                    number_format=src.number_format or "General",
                    fill_color=fill_color,
                    font_color=font_color,
                    bold=bool(font and font.bold),
                    italic=bool(font and font.italic),
                    underline=bool(
                        font and font.underline and font.underline != "none"
                    ),
                    strike=bool(font and font.strike),
                    font_size=float(font.size) if font and font.size else None,
                )
            )
        grid.append(row_cells)

    # Handle merged cells: propagate top-left value/format to the whole range,
    # and mark bounds so downstream logic knows this was a merge.
    for merged_range in ws.merged_cells.ranges:
        min_r, min_c = merged_range.min_row, merged_range.min_col
        max_r, max_c = merged_range.max_row, merged_range.max_col
        anchor = grid[min_r - 1][min_c - 1]
        anchor.is_merged_anchor = True
        anchor.merged_range = str(merged_range)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                cell = grid[r - 1][c - 1]
                cell.value = anchor.value
                cell.number_format = anchor.number_format
                cell.merged_range = str(merged_range)
                # The anchor's formatting paints across the whole merge.
                cell.fill_color = anchor.fill_color
                cell.font_color = anchor.font_color
                cell.bold = anchor.bold
                cell.italic = anchor.italic
                cell.underline = anchor.underline
                cell.strike = anchor.strike
                cell.font_size = anchor.font_size

    return Sheet(grid, n_rows, n_cols, name=ws.title)
