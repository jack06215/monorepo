"""
Unit tests for the SheetCompressor reimplementation.
Uses only the standard library (unittest) so it runs with no extra installs.

Run with:  python3 -m unittest test_sheet_compressor -v
       or: python3 test_sheet_compressor.py
"""

import datetime
import os
import unittest

import openpyxl

from packages.spreadsheet_llm.data_format_aggregation import (
    aggregate_by_data_format,
    infer_data_type,
)
from packages.spreadsheet_llm.inverted_index import (
    build_inverted_index,
    render_paper_style,
)
from packages.spreadsheet_llm.sheet_compressor import (
    compress_sheet,
    count_tokens,
    vanilla_encode,
)
from packages.spreadsheet_llm.sheet_model import Cell, load_xlsx
from packages.spreadsheet_llm.structural_anchors import extract_structural_anchors

TMP_PATH = "/tmp/_sc_test.xlsx"


def _make_simple_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 90
    ws["A3"] = "Bob"
    ws["B3"] = 85
    wb.save(TMP_PATH)
    return TMP_PATH


class SheetModelTests(unittest.TestCase):
    def tearDown(self):
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_load_xlsx_basic_shape(self):
        path = _make_simple_xlsx()
        sheet = load_xlsx(path)
        self.assertEqual(sheet.n_rows, 3)
        self.assertEqual(sheet.n_cols, 2)
        self.assertEqual(sheet.get(1, 1).value, "Name")
        self.assertEqual(sheet.get(2, 2).value, 90)

    def test_cell_address(self):
        c = Cell(row=5, col=3, value="x", number_format="General")
        self.assertEqual(c.address, "C5")

    def test_cell_is_empty(self):
        self.assertTrue(
            Cell(row=1, col=1, value=None, number_format="General").is_empty
        )
        self.assertTrue(
            Cell(row=1, col=1, value="  ", number_format="General").is_empty
        )
        self.assertFalse(Cell(row=1, col=1, value=0, number_format="General").is_empty)

    def test_merged_cell_value_propagates(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws.merge_cells("A1:C1")
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        self.assertEqual(sheet.get(1, 1).value, "Header")
        self.assertEqual(sheet.get(1, 2).value, "Header")
        self.assertEqual(sheet.get(1, 3).value, "Header")
        self.assertTrue(sheet.get(1, 1).is_merged_anchor)
        self.assertFalse(sheet.get(1, 2).is_merged_anchor)


class StructuralAnchorTests(unittest.TestCase):
    def tearDown(self):
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_extraction_keeps_headers_drops_homogeneous_middle(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "ID"
        ws["B1"] = "Value"
        for i in range(2, 102):  # 100 homogeneous data rows
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 1.5)
        wb.save(TMP_PATH)

        sheet = load_xlsx(TMP_PATH)
        result = extract_structural_anchors(sheet, k=4)

        self.assertIn(1, result.kept_rows)  # header
        self.assertIn(2, result.kept_rows)  # near boundary
        self.assertNotIn(50, result.kept_rows)  # deep homogeneous middle -> dropped
        self.assertLess(len(result.kept_rows), sheet.n_rows)

    def test_extraction_remap_is_contiguous(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "ID"
        for i in range(2, 60):
            ws.cell(row=i, column=1, value=i)
        wb.save(TMP_PATH)

        sheet = load_xlsx(TMP_PATH)
        result = extract_structural_anchors(sheet, k=3)
        new_rows = sorted(result.row_map.values())
        self.assertEqual(new_rows, list(range(1, len(new_rows) + 1)))

    def test_extraction_never_crashes_on_tiny_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "x"
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        result = extract_structural_anchors(sheet, k=4)
        self.assertGreaterEqual(result.sheet.n_rows, 1)
        self.assertGreaterEqual(result.sheet.n_cols, 1)


class DataFormatAggregationTests(unittest.TestCase):
    def tearDown(self):
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_infer_data_type_int(self):
        c = Cell(row=1, col=1, value=42, number_format="General")
        self.assertEqual(infer_data_type(c), "IntNum")

    def test_infer_data_type_float(self):
        c = Cell(row=1, col=1, value=3.14, number_format="General")
        self.assertEqual(infer_data_type(c), "FloatNum")

    def test_infer_data_type_date_by_nfs(self):
        c = Cell(
            row=1, col=1, value=datetime.date(2024, 1, 1), number_format="yyyy-mm-dd"
        )
        self.assertEqual(infer_data_type(c), "DateData")

    def test_infer_data_type_currency_by_nfs(self):
        c = Cell(row=1, col=1, value=19.99, number_format="$#,##0.00")
        self.assertEqual(infer_data_type(c), "CurrencyData")

    def test_infer_data_type_percentage_by_nfs(self):
        c = Cell(row=1, col=1, value=0.5, number_format="0.0%")
        self.assertEqual(infer_data_type(c), "PercentageData")

    def test_infer_data_type_email(self):
        c = Cell(row=1, col=1, value="user@example.com", number_format="General")
        self.assertEqual(infer_data_type(c), "EmailData")

    def test_infer_data_type_string(self):
        c = Cell(row=1, col=1, value="Hello World", number_format="General")
        self.assertEqual(infer_data_type(c), "String")

    def test_infer_data_type_empty(self):
        c = Cell(row=1, col=1, value=None, number_format="General")
        self.assertEqual(infer_data_type(c), "Empty")

    def test_aggregation_merges_adjacent_same_type_run(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 11):
            ws.cell(row=i, column=1, value=i)
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        regions, _ = aggregate_by_data_format(sheet)
        int_regions = [r for r in regions if r.data_type == "IntNum"]
        self.assertEqual(len(int_regions), 1)
        self.assertEqual(int_regions[0].top, 1)
        self.assertEqual(int_regions[0].bottom, 10)

    def test_aggregation_does_not_merge_across_type_boundary(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
        for i in range(6, 11):
            ws.cell(row=i, column=1, value=f"text{i}")
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        regions, _ = aggregate_by_data_format(sheet)
        int_regions = [r for r in regions if r.data_type == "IntNum"]
        string_regions = [r for r in regions if r.data_type == "String"]
        self.assertEqual(len(int_regions), 1)
        self.assertEqual(int_regions[0].bottom, 5)
        self.assertEqual(len(string_regions), 5)


class InvertedIndexTests(unittest.TestCase):
    def tearDown(self):
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_inverted_index_skips_empty_and_dedupes_values(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Yes"
        ws["A2"] = "Yes"
        ws["B1"] = None
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        regions, _ = aggregate_by_data_format(sheet)
        index = build_inverted_index(sheet, regions)
        self.assertIn("Yes", index)
        self.assertEqual(set(index["Yes"]), {"A1", "A2"})
        self.assertTrue(all(v != "" for v in index.keys()))

    def test_render_paper_style_format(self):
        index = {"Year": ["A1"], "IntNum": ["A2:B3"]}
        text = render_paper_style(index)
        self.assertEqual(text, "(Year|A1)(IntNum|A2:B3)")


class FullPipelineTests(unittest.TestCase):
    def tearDown(self):
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_vanilla_encode_includes_empty_cells(self):
        path = _make_simple_xlsx()
        sheet = load_xlsx(path)
        text = vanilla_encode(sheet)
        self.assertGreaterEqual(text.count(","), 6)

    def test_compression_ratio_is_positive_and_greater_than_one_on_redundant_sheet(
        self,
    ):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "ID"
        ws["B1"] = "Date"
        for i in range(2, 60):
            ws.cell(row=i, column=1, value=i)
            d = ws.cell(row=i, column=2, value=datetime.date(2024, 1, 1))
            d.number_format = "yyyy-mm-dd"
        wb.save(TMP_PATH)
        sheet = load_xlsx(TMP_PATH)
        report = compress_sheet(sheet, k=4)
        self.assertGreater(report.compression_ratio, 1.0)
        self.assertLess(report.compressed_tokens, report.vanilla_tokens)

    def test_pipeline_runs_without_aggregation_ablation(self):
        path = _make_simple_xlsx()
        sheet = load_xlsx(path)
        report = compress_sheet(sheet, use_aggregation=False)
        self.assertTrue(report.compressed_text)
        self.assertNotIn("IntNum", report.compressed_text)
        self.assertTrue(
            "90" in report.compressed_text or "85" in report.compressed_text
        )

    def test_pipeline_content_preserved_end_to_end(self):
        path = _make_simple_xlsx()
        sheet = load_xlsx(path)
        report = compress_sheet(sheet, use_aggregation=True)
        for name in ["Name", "Score", "Alice", "Bob"]:
            self.assertIn(name, report.compressed_text)

    def test_count_tokens_nonzero_for_nonempty_text(self):
        self.assertGreater(count_tokens("hello world this is a test"), 0)

    def test_count_tokens_zero_for_empty_text(self):
        self.assertEqual(count_tokens(""), 0)


if __name__ == "__main__":
    unittest.main(verbosity=2)
