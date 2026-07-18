"""
Generates the synthetic .xlsx fixtures under sample_data/:

- test_sheet.xlsx: a multi-table sheet that mimics the kind of sheet shown
  in the paper's Figure 2: a title row, two tables stacked with a gap, a
  long homogeneous data region, mixed number formats (dates, currency,
  percentages, plain ints), and an email column -- so we can exercise all
  three SheetCompressor modules meaningfully.
- big_sheet.xlsx: a 303-row single-table transaction log whose homogeneous
  body is long enough to trigger llm_context_encoder's row-sampling and
  gap-annotation path.
"""

import datetime
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font

SAMPLE_DATA_DIR = pathlib.Path(__file__).resolve().parent / "sample_data"


def build(output_dir: pathlib.Path = SAMPLE_DATA_DIR) -> pathlib.Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Title (row 1) -- should NOT be treated as a table header/body
    ws["B1"] = "Quarterly Sales Report - Region: North America"
    ws["B1"].font = Font(bold=True, size=14)

    # ---- Table 1: header at row 3, data rows 4-53 (50 homogeneous rows) ----
    headers1 = ["Product", "Date", "UnitsSold", "UnitPrice", "Revenue", "Contact"]
    for j, h in enumerate(headers1):
        ws.cell(row=3, column=2 + j, value=h).font = Font(bold=True)

    products = ["Widget", "Gadget", "Gizmo", "Doohickey", "Thingamajig"]
    for i in range(50):
        r = 4 + i
        ws.cell(row=r, column=2, value=products[i % len(products)])
        d = ws.cell(
            row=r,
            column=3,
            value=datetime.date(2024, 1, 1) + datetime.timedelta(days=i),
        )
        d.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=4, value=10 + i)  # UnitsSold (int)
        price = ws.cell(row=r, column=5, value=9.99 + (i % 5))
        price.number_format = "$#,##0.00"
        rev = ws.cell(row=r, column=6, value=(10 + i) * (9.99 + (i % 5)))
        rev.number_format = "$#,##0.00"
        email = ws.cell(row=r, column=7, value=f"buyer{i}@example.com")

    # Gap rows 54-56 (empty) to separate tables

    # ---- Table 2: header at row 57, data rows 58-67 (small, different shape) ----
    headers2 = ["Region", "Quota", "Achieved", "PctOfQuota"]
    for j, h in enumerate(headers2):
        ws.cell(row=57, column=2 + j, value=h).font = Font(bold=True)

    regions = [
        "East",
        "West",
        "North",
        "South",
        "Central",
        "NE",
        "NW",
        "SE",
        "SW",
        "Mid",
    ]
    for i in range(10):
        r = 58 + i
        ws.cell(row=r, column=2, value=regions[i])
        quota = ws.cell(row=r, column=3, value=100000 + i * 1000)
        achieved = ws.cell(row=r, column=4, value=90000 + i * 1500)
        pct = ws.cell(row=r, column=5, value=(90000 + i * 1500) / (100000 + i * 1000))
        pct.number_format = "0.0%"

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "test_sheet.xlsx"
    wb.save(path)
    print(f"Saved {path}:", ws.max_row, "rows x", ws.max_column, "cols")
    return path


def build_big_sheet(
    output_dir: pathlib.Path = SAMPLE_DATA_DIR, n_data_rows: int = 300
) -> pathlib.Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Title (row 1), header (row 3), then one long homogeneous body: exactly
    # the shape where Module 1 alone silently drops the middle and
    # llm_context_encoder must annotate the gap.
    ws["A1"] = "Transaction Log 2024"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["TxnID", "Date", "Merchant", "Category", "Amount"]
    for j, h in enumerate(headers):
        ws.cell(row=3, column=1 + j, value=h).font = Font(bold=True)

    merchants = ["Acme Corp", "Globex", "Initech", "Umbrella", "Stark Industries"]
    categories = ["Office", "Travel", "Software", "Hardware", "Meals"]
    for i in range(n_data_rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=f"TXN-{i:05d}")
        d = ws.cell(
            row=r,
            column=2,
            value=datetime.date(2024, 1, 1) + datetime.timedelta(days=i % 365),
        )
        d.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3, value=merchants[i % len(merchants)])
        ws.cell(row=r, column=4, value=categories[i % len(categories)])
        amount = ws.cell(row=r, column=5, value=10.5 + (i % 90) * 3.25)
        amount.number_format = "$#,##0.00"

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "big_sheet.xlsx"
    wb.save(path)
    print(f"Saved {path}:", ws.max_row, "rows x", ws.max_column, "cols")
    return path


if __name__ == "__main__":
    build()
    build_big_sheet()
