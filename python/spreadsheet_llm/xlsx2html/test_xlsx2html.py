"""Unit tests for the xlsx2html package (openpyxl-native, style-preserving
Excel -> HTML conversion).

Run with:  poetry run python -m pytest python/spreadsheet_llm/xlsx2html/
"""

import datetime
import os
import re
import unittest

import openpyxl
import openpyxl.drawing.image  # type: ignore[import-not-found]
from openpyxl.comments import Comment  # type: ignore[import-not-found]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from PIL import Image

from python.spreadsheet_llm.xlsx2html.parser import Xlsx2Html, _render_value
from python.spreadsheet_llm.xlsx2html.style_util import (DEFAULT_THEME_PALETTE,
                                                         _apply_tint,
                                                         extract_theme_palette,
                                                         resolve_color)
from python.spreadsheet_llm.xlsx2html.types import ParsedWorksheet

TMP_PATH = "/tmp/_xlsx2html_test.xlsx"
TMP_PNG = "/tmp/_xlsx2html_test.png"

XlsxImage = openpyxl.drawing.image.Image


def _save(wb: openpyxl.Workbook) -> str:
    wb.save(TMP_PATH)
    return TMP_PATH


def _parse_first(wb: openpyxl.Workbook) -> ParsedWorksheet:
    return Xlsx2Html(filename=_save(wb)).parse()[0]


class ResolveColorTests(unittest.TestCase):
    def test_rgb_color_drops_alpha(self) -> None:
        self.assertEqual(
            resolve_color(Color(rgb="FF12AB34"), DEFAULT_THEME_PALETTE), "#12AB34"
        )

    def test_indexed_color_resolves_via_legacy_palette(self) -> None:
        # COLOR_INDEX[2] is red in the legacy palette.
        self.assertEqual(
            resolve_color(Color(indexed=2), DEFAULT_THEME_PALETTE), "#FF0000"
        )

    def test_indexed_system_auto_colors_resolve_to_none(self) -> None:
        self.assertIsNone(resolve_color(Color(indexed=64), DEFAULT_THEME_PALETTE))

    def test_theme_color_resolves_via_palette(self) -> None:
        self.assertEqual(
            resolve_color(Color(theme=4, tint=0.0), DEFAULT_THEME_PALETTE),
            f"#{DEFAULT_THEME_PALETTE[4]}",
        )

    def test_positive_tint_lightens(self) -> None:
        base = "4472C4"
        tinted = _apply_tint(base, 0.4)
        self.assertNotEqual(tinted, base)
        # Every channel moves toward white.
        for i in (0, 2, 4):
            self.assertGreaterEqual(int(tinted[i : i + 2], 16), int(base[i : i + 2], 16))

    def test_auto_color_resolves_to_none(self) -> None:
        self.assertIsNone(resolve_color(Color(auto=True), DEFAULT_THEME_PALETTE))
        self.assertIsNone(resolve_color(None, DEFAULT_THEME_PALETTE))


class RenderValueTests(unittest.TestCase):
    def test_midnight_datetime_with_date_format_renders_iso_date(self) -> None:
        value = datetime.datetime(2024, 3, 5)
        self.assertEqual(_render_value(value, "yyyy-mm-dd"), "2024-03-05")

    def test_datetime_with_time_format_keeps_time(self) -> None:
        value = datetime.datetime(2024, 3, 5, 14, 30)
        self.assertEqual(_render_value(value, "yyyy-mm-dd hh:mm"), "2024-03-05 14:30:00")

    def test_percentage_format_renders_percent(self) -> None:
        self.assertEqual(_render_value(0.9, "0.0%"), "90%")
        self.assertEqual(_render_value(0.905940594059406, "0.0%"), "90.5941%")

    def test_whole_float_renders_without_decimal(self) -> None:
        self.assertEqual(_render_value(17.0, "General"), "17")

    def test_bool_and_none(self) -> None:
        self.assertEqual(_render_value(True, "General"), "TRUE")
        self.assertEqual(_render_value(None, "General"), "")


class Xlsx2HtmlTests(unittest.TestCase):
    def tearDown(self) -> None:
        for path in (TMP_PATH, TMP_PNG):
            if os.path.exists(path):
                os.remove(path)

    def test_styled_cell_gets_css_class_with_expected_props(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Header"
        ws["A1"].font = Font(bold=True, color=Color(rgb="FFFF0000"))
        ws["A1"].fill = PatternFill(patternType="solid", fgColor=Color(rgb="FFFFFF00"))
        parsed = _parse_first(wb)
        match = re.search(r"<td class=\"(s\d+)\"", parsed.html)
        self.assertIsNotNone(match)
        assert match is not None
        rule = re.search(rf"\.{match.group(1)}\{{([^}}]*)\}}", parsed.html)
        assert rule is not None
        self.assertIn("font-weight:bold", rule.group(1))
        self.assertIn("color:#FF0000", rule.group(1))
        self.assertIn("background:#FFFF00", rule.group(1))

    def test_identical_styles_share_one_class_and_plain_cells_have_none(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "one"
        ws["B1"] = "two"
        bold = Font(bold=True)
        ws["A1"].font = bold
        ws["B1"].font = bold
        ws["A2"] = "plain"
        parsed = _parse_first(wb)
        classes = re.findall(r'class="(s\d+)"', parsed.html)
        self.assertEqual(classes, ["s1", "s1"])
        self.assertEqual(parsed.html.count("<style>"), 1)
        self.assertIn("<td>plain</td>", parsed.html)

    def test_theme_color_with_tint_resolves_to_hex(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "themed"
        ws["A1"].font = Font(color=Color(theme=4, tint=0.4))
        path = _save(wb)
        converter = Xlsx2Html(filename=path)
        palette = extract_theme_palette(converter._workbook)
        expected = f"#{_apply_tint(palette[4], 0.4)}"
        self.assertIn(f"color:{expected}", converter.parse()[0].html)

    def test_merged_range_becomes_colspan_rowspan(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Title"
        ws.merge_cells("A1:C2")
        ws["A3"] = "below"
        ws["C3"] = "corner"
        parsed = _parse_first(wb)
        self.assertIn('rowspan="2"', parsed.html)
        self.assertIn('colspan="3"', parsed.html)
        # 1 anchor cell in rows 1-2 (5 covered cells skipped) + 3 cells in row 3.
        self.assertEqual(parsed.html.count("<td"), 4)

    def test_number_format_exposed_as_data_nf(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = datetime.date(2024, 1, 15)  # type: ignore[assignment]
        ws["A1"].number_format = "yyyy-mm-dd"
        ws["B1"] = 0.9
        ws["B1"].number_format = "0.0%"
        ws["C1"] = "plain"
        parsed = _parse_first(wb)
        self.assertIn('data-nf="yyyy-mm-dd">2024-01-15<', parsed.html)
        self.assertIn(">90%<", parsed.html)
        self.assertIn("<td>plain</td>", parsed.html)

    def test_hyperlink_comment_and_escaping(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "docs"
        ws["A1"].hyperlink = "https://example.com/?a=1&b=2"
        ws["B1"] = "a < b & c"
        ws["B1"].comment = Comment("check <this>", "author")
        parsed = _parse_first(wb)
        self.assertIn('<a href="https://example.com/?a=1&amp;b=2">docs</a>', parsed.html)
        self.assertIn(">a &lt; b &amp; c<", parsed.html)
        self.assertIn('title="check &lt;this&gt;"', parsed.html)

    def test_hidden_row_and_column_are_marked(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "visible"
        ws["B1"] = "hidden col"
        ws["A2"] = "hidden row"
        ws.row_dimensions[2].hidden = True
        ws.column_dimensions["B"].hidden = True
        parsed = _parse_first(wb)
        self.assertIn('<tr class="hidden"><td>hidden row</td>', parsed.html)
        self.assertIn('data-hidden-cols="B"', parsed.html)

    def test_image_is_placeholder_not_base64_and_survives_past_column_z(self) -> None:
        Image.new("RGB", (2, 2), "red").save(TMP_PNG)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "has image sheet"
        ws.add_image(XlsxImage(TMP_PNG), "AB2")
        parsed = _parse_first(wb)
        self.assertIn("<td>[image]</td>", parsed.html)
        self.assertEqual(len(parsed.base64_encoded_images), 1)
        # The base64 payload must stay off the HTML (token cost).
        self.assertNotIn(parsed.base64_encoded_images[0][:16], parsed.html)

    def test_trailing_empty_rows_and_cols_trimmed(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "x"
        ws["B2"] = "y"
        ws["E9"] = None  # type: ignore[assignment]  # touched but empty -> must not count
        parsed = _parse_first(wb)
        self.assertEqual(parsed.html.count("<tr"), 2)
        self.assertEqual(parsed.html.count("<td"), 4)

    def test_multi_sheet_workbook_yields_fragment_per_sheet(self) -> None:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1["A1"] = "one"
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "two"
        parsed = Xlsx2Html(filename=_save(wb)).parse()
        self.assertEqual([p.worksheet_name for p in parsed], ["First", "Second"])
        self.assertIn('data-sheet="First"', parsed[0].html)
        self.assertIn('data-sheet="Second"', parsed[1].html)


if __name__ == "__main__":
    unittest.main(verbosity=2)
