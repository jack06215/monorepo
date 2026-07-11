"""
Unit tests for llm_context_encoder (Module 1 + row sampling + Module 2,
no data-format aggregation).

Run with:  python3 -m unittest test_llm_context_encoder -v
       or: python3 -m pytest python/spreadsheet_llm/test_llm_context_encoder.py
"""

import datetime
import os
import unittest

import openpyxl

from python.spreadsheet_llm.llm_context_encoder import (_find_module1_gaps,
                                                        apply_row_sampling,
                                                        build_llm_context,
                                                        literal_regions,
                                                        sample_long_row_runs)
from python.spreadsheet_llm.sheet_model import Cell, Sheet, load_xlsx

TMP_PATH = "/tmp/_llm_ctx_test.xlsx"


def _make_sheet(n_rows: int, n_cols: int = 1) -> Sheet:
    grid = [
        [
            Cell(row=r, col=c, value=f"v{r}_{c}", number_format="General")
            for c in range(1, n_cols + 1)
        ]
        for r in range(1, n_rows + 1)
    ]
    return Sheet(grid, n_rows, n_cols)


def _make_simple_xlsx() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 90
    ws["A3"] = "Bob"
    ws["B3"] = 85
    wb.save(TMP_PATH)
    return TMP_PATH


def _make_big_xlsx(n_data_rows: int = 300) -> str:
    """Single table: header at row 1, homogeneous body rows 2..n_data_rows+1.

    The shape where Module 1 alone silently drops a huge middle -- the bug
    _find_module1_gaps exists to annotate.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for j, h in enumerate(["TxnID", "Date", "Amount"]):
        ws.cell(row=1, column=1 + j, value=h)
    for i in range(n_data_rows):
        r = 2 + i
        ws.cell(row=r, column=1, value=f"TXN-{i:05d}")
        d = ws.cell(
            row=r,
            column=2,
            value=datetime.date(2024, 1, 1) + datetime.timedelta(days=i % 365),
        )
        d.number_format = "yyyy-mm-dd"
        amount = ws.cell(row=r, column=3, value=10.5 + (i % 90) * 3.25)
        amount.number_format = "$#,##0.00"
    wb.save(TMP_PATH)
    return TMP_PATH


class SampleLongRowRunsTests(unittest.TestCase):
    def test_short_run_kept_whole(self) -> None:
        sheet = _make_sheet(10)
        result = sample_long_row_runs(
            sheet, {1}, sample_rows=5, min_run_to_sample=15
        )
        self.assertEqual(result.kept_rows, list(range(1, 11)))
        self.assertEqual(result.omitted_runs, [])

    def test_run_exactly_min_length_not_sampled(self) -> None:
        # Run of rows 2..16 is exactly 15 rows -- threshold is strict (>).
        sheet = _make_sheet(16)
        result = sample_long_row_runs(
            sheet, {1}, sample_rows=5, min_run_to_sample=15
        )
        self.assertEqual(result.kept_rows, list(range(1, 17)))
        self.assertEqual(result.omitted_runs, [])

    def test_long_run_head_tail_sampled(self) -> None:
        # Run of rows 2..30 (29 rows) -> keep 2-6 and 26-30, omit 7-25.
        sheet = _make_sheet(30)
        result = sample_long_row_runs(
            sheet, {1}, sample_rows=5, min_run_to_sample=15
        )
        self.assertEqual(result.omitted_runs, [(7, 25)])
        self.assertEqual(
            result.kept_rows, [1] + list(range(2, 7)) + list(range(26, 31))
        )

    def test_run_between_two_anchors_sampled(self) -> None:
        # Anchors at 1 and 30 bracket a 28-row body run (2..29).
        sheet = _make_sheet(30)
        result = sample_long_row_runs(
            sheet, {1, 30}, sample_rows=5, min_run_to_sample=15
        )
        self.assertEqual(result.omitted_runs, [(7, 24)])
        self.assertIn(1, result.kept_rows)
        self.assertIn(30, result.kept_rows)

    def test_anchor_rows_always_kept(self) -> None:
        sheet = _make_sheet(40)
        anchors = {1, 20, 40}
        result = sample_long_row_runs(
            sheet, anchors, sample_rows=3, min_run_to_sample=5
        )
        for a in anchors:
            self.assertIn(a, result.kept_rows)


class ApplyRowSamplingTests(unittest.TestCase):
    def test_rows_remapped_contiguously(self) -> None:
        sheet = _make_sheet(12, n_cols=2)
        sampling = sample_long_row_runs(
            sheet, {1}, sample_rows=2, min_run_to_sample=4
        )
        new_sheet, row_map = apply_row_sampling(sheet, sampling)
        self.assertEqual(new_sheet.n_rows, len(sampling.kept_rows))
        self.assertEqual(
            sorted(row_map.values()), list(range(1, len(sampling.kept_rows) + 1))
        )
        for r in range(1, new_sheet.n_rows + 1):
            self.assertEqual(new_sheet.get(r, 1).row, r)

    def test_values_preserved_after_remap(self) -> None:
        sheet = _make_sheet(12, n_cols=2)
        sampling = sample_long_row_runs(
            sheet, {1}, sample_rows=2, min_run_to_sample=4
        )
        new_sheet, row_map = apply_row_sampling(sheet, sampling)
        for orig, new in row_map.items():
            for c in (1, 2):
                self.assertEqual(new_sheet.get(new, c).value, sheet.get(orig, c).value)


class FindModule1GapsTests(unittest.TestCase):
    def test_no_gap_for_contiguous_rows(self) -> None:
        self.assertEqual(_find_module1_gaps([1, 2, 3, 4]), [])

    def test_gaps_detected(self) -> None:
        self.assertEqual(
            _find_module1_gaps([1, 2, 3, 10, 11, 20]), [(4, 9), (12, 19)]
        )


class LiteralRegionsTests(unittest.TestCase):
    def test_one_region_per_cell_no_aggregation(self) -> None:
        grid = [
            [
                Cell(row=1, col=1, value="Name", number_format="General"),
                Cell(row=1, col=2, value=42, number_format="General"),
            ],
            [
                Cell(row=2, col=1, value=None, number_format="General"),
                Cell(row=2, col=2, value=43, number_format="General"),
            ],
        ]
        sheet = Sheet(grid, 2, 2)
        regions = literal_regions(sheet)
        self.assertEqual(len(regions), 4)  # one per cell, ints NOT merged
        for reg in regions:
            self.assertEqual((reg.top, reg.left), (reg.bottom, reg.right))
        dtypes = {(reg.top, reg.left): reg.data_type for reg in regions}
        self.assertEqual(dtypes[(2, 1)], "Empty")
        self.assertEqual(dtypes[(1, 2)], "String")  # literal, no "IntNum"


class BuildLLMContextTests(unittest.TestCase):
    def tearDown(self) -> None:
        if os.path.exists(TMP_PATH):
            os.remove(TMP_PATH)

    def test_small_sheet_complete_and_unannotated(self) -> None:
        path = _make_simple_xlsx()
        result = build_llm_context(path)
        self.assertNotIn("[NOTE:", result.compressed_text)
        self.assertEqual(result.omitted_row_runs, [])
        for val in ["Name", "Score", "Alice", "Bob", "90", "85"]:
            self.assertIn(val, result.compressed_text)

    def test_big_sheet_gap_is_annotated(self) -> None:
        path = _make_big_xlsx(300)
        result = build_llm_context(path)
        self.assertIn("[NOTE:", result.compressed_text)
        self.assertIn("omitted", result.compressed_text)
        self.assertTrue(result.omitted_row_runs)
        # Header and both ends of the body survive verbatim.
        for val in ["TxnID", "Date", "Amount", "TXN-00000", "TXN-00299"]:
            self.assertIn(val, result.compressed_text)

    def test_middle_rows_dropped_but_accounted_for(self) -> None:
        path = _make_big_xlsx(300)
        result = build_llm_context(path)
        # TXN-00150 lives at original row 152 -- deep in the homogeneous
        # middle, so it must be gone from the text but covered by a NOTE run.
        self.assertNotIn("TXN-00150", result.compressed_text)
        self.assertTrue(
            any(a <= 152 <= b for a, b in result.omitted_row_runs),
            f"row 152 not covered by any omitted run: {result.omitted_row_runs}",
        )

    def test_no_silent_row_loss(self) -> None:
        # Regression test: Module 1 alone can drop a huge homogeneous middle
        # with no annotation. Every original row must be either kept or
        # covered by an omitted-run annotation.
        path = _make_big_xlsx(300)
        result = build_llm_context(path)
        omitted = sum(b - a + 1 for a, b in result.omitted_row_runs)
        self.assertEqual(result.after_sampling_rows + omitted, result.original_rows)

    def test_row_sampling_path_also_annotates(self) -> None:
        # A large k makes Module 1 keep a wide anchor neighborhood, so the
        # leftover long body run must be thinned by row sampling instead.
        path = _make_big_xlsx(100)
        result = build_llm_context(path, k=50, sample_rows=5, min_run_to_sample=15)
        # Module 1 kept everything (k covers the whole sheet), so any
        # thinning here is row sampling's doing.
        self.assertEqual(result.after_module1_rows, result.original_rows)
        self.assertLess(result.after_sampling_rows, result.after_module1_rows)
        self.assertIn("[NOTE:", result.compressed_text)
        omitted = sum(b - a + 1 for a, b in result.omitted_row_runs)
        self.assertEqual(result.after_sampling_rows + omitted, result.original_rows)

    def test_no_type_labels_real_values_only(self) -> None:
        path = _make_big_xlsx(300)
        result = build_llm_context(path)
        for label in ["IntNum", "FloatNum", "DateData", "CurrencyData", "EmailData"]:
            self.assertNotIn(label, result.compressed_text)
        self.assertIn("2024-01-01", str(result.compressed_text))

    def test_compression_ratio_positive_on_big_sheet(self) -> None:
        path = _make_big_xlsx(300)
        result = build_llm_context(path)
        self.assertGreater(result.compression_ratio, 1.0)
        self.assertLess(result.compressed_tokens, result.vanilla_tokens)

    def test_json_output_style(self) -> None:
        path = _make_simple_xlsx()
        result = build_llm_context(path, output_style="json")
        self.assertTrue(result.compressed_text.startswith("{"))
        self.assertIn('"Alice"', result.compressed_text)
        self.assertNotIn("(Alice|", result.compressed_text)

    def test_sheet_name_selection(self) -> None:
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1["A1"] = "first-sheet-value"
        ws2 = wb.create_sheet("Second")
        ws2["A1"] = "second-sheet-value"
        wb.save(TMP_PATH)
        result = build_llm_context(TMP_PATH, sheet_name="Second")
        self.assertIn("second-sheet-value", result.compressed_text)
        self.assertNotIn("first-sheet-value", result.compressed_text)


if __name__ == "__main__":
    unittest.main(verbosity=2)
