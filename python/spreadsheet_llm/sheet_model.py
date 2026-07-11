"""
sheet_model.py

Loads a real .xlsx worksheet (via openpyxl) into a simple in-memory grid of
Cell objects that the rest of SheetCompressor operates on.

This is the "Section 3.1 vanilla" data model: S = {Cell_i,j} for i in m, j in n.
"""

from dataclasses import dataclass
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class Cell:
    row: int  # 1-indexed row
    col: int  # 1-indexed column
    value: object  # raw python value (str/int/float/date/None)
    number_format: str  # Excel's NFS, e.g. "yyyy-mm-dd", "#,##0.00", "General"
    is_merged_anchor: bool = (
        False  # True if this is the top-left cell of a merged range
    )
    merged_range: Optional[str] = None

    @property
    def address(self) -> str:
        return f"{get_column_letter(self.col)}{self.row}"

    @property
    def is_empty(self) -> bool:
        return self.value is None or (
            isinstance(self.value, str) and self.value.strip() == ""
        )


class Sheet:
    """A dense m x n grid of Cell objects, 1-indexed to match Excel addressing."""

    def __init__(
        self, cells: List[List[Cell]], n_rows: int, n_cols: int, name: str = "Sheet1"
    ):
        self.cells = cells  # cells[row-1][col-1] -> Cell
        self.n_rows = n_rows
        self.n_cols = n_cols
        self.name = name

    def get(self, row: int, col: int) -> Cell:
        return self.cells[row - 1][col - 1]

    def __repr__(self):
        return f"Sheet(name={self.name!r}, rows={self.n_rows}, cols={self.n_cols})"


def load_xlsx(path: str, sheet_name: Optional[str] = None) -> Sheet:
    """
    Load a worksheet from an xlsx file into a Sheet of Cells.

    - Uses openpyxl with data_only=True so formulas resolve to their last
      calculated values (falls back to formula string if no cached value).
    - Fills in every cell in the used range, including empty ones, so the
      grid is dense (required for structural-anchor / aggregation steps).
    - Propagates merged-cell values to every cell in the merged range, since
      openpyxl only stores the value on the top-left cell of a merge; this
      matches how a person visually reads a merged header.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0

    # Pre-fetch number formats + values in one pass
    grid: List[List[Cell]] = []
    for r in range(1, n_rows + 1):
        row_cells = []
        for c in range(1, n_cols + 1):
            src = ws.cell(row=r, column=c)
            row_cells.append(
                Cell(
                    row=r,
                    col=c,
                    value=src.value,
                    number_format=src.number_format or "General",
                )
            )
        grid.append(row_cells)

    # Handle merged cells: propagate top-left value/format to the whole range,
    # and mark bounds so downstream logic knows this was a merge.
    for merged_range in ws.merged_cells.ranges:
        min_r, min_c = merged_range.min_row, merged_range.min_col
        max_r, max_c = merged_range.max_row, merged_range.max_col
        anchor = grid[min_r - 1][min_c - 1]
        anchor.is_merged_anchor = True
        anchor.merged_range = str(merged_range)
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                cell = grid[r - 1][c - 1]
                cell.value = anchor.value
                cell.number_format = anchor.number_format
                cell.merged_range = str(merged_range)

    return Sheet(grid, n_rows, n_cols, name=ws.title)
