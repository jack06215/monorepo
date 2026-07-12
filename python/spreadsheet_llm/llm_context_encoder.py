"""
llm_context_encoder.py

A leaner SheetCompressor variant for an ingestion pipeline where an LLM
reads a compressed spreadsheet and generates a *description* of it (which
then gets indexed into Elasticsearch alongside/instead of the raw data).

Runs only:
  - Module 1 (structural-anchor extraction)  -> cheap way to find table
    boundaries/headers and drop long homogeneous stretches on large sheets
  - Module 2 (inverted-index translation)     -> compact, lossless-for-what-
    survives serialization: {value: address_or_range}, empty cells dropped

Deliberately SKIPS Module 3 (data-format-aware aggregation). That module
replaces real values with type labels like "IntNum"/"DateData", which is
exactly what you don't want when the LLM's job is to write a description
containing real specifics (product names, actual date ranges, example
values) rather than a structurally-accurate-but-content-free summary.

To keep large sheets within context even without Module 3's aggressive
numeric collapsing, this script adds a row-sampling strategy: after
Module 1 narrows down to the structurally important rows, if a detected
table's *body* is still long, keep the first N and last N body rows
verbatim (headers are always kept in full) and mark the gap explicitly,
so the LLM sees real example rows from both ends of each table instead of
type placeholders.

Visual formatting is preserved too: cell fill color, font color, and font
style (bold/italic/underline/strikethrough, outlier font sizes) are loaded
from the workbook, carried through extraction/sampling, and emitted as a
compact [FORMATTING] index (style -> cell ranges) after the value tuples,
since header bands, color-coded cells, and bold totals carry table
semantics that values alone lose.

Usage:
    python3 llm_context_encoder.py my_file.xlsx
    python3 llm_context_encoder.py my_file.xlsx --k 4 --sample-rows 5
    python3 llm_context_encoder.py my_file.xlsx --format json
    python3 llm_context_encoder.py my_file.xlsx --no-gap-notes
    python3 llm_context_encoder.py my_file.xlsx --no-row-sampling
    python3 llm_context_encoder.py my_file.xlsx --no-module1
"""

import argparse
from collections import Counter
from dataclasses import dataclass, replace
from typing import Dict, List, Optional, Set, Tuple

from openpyxl.utils import get_column_letter

from python.spreadsheet_llm.data_format_aggregation import (
    AggregatedRegion,  # reused as a plain container, no aggregation logic invoked
)
from python.spreadsheet_llm.inverted_index import (
    build_inverted_index,
    render_json_like,
    render_paper_style,
)
from python.spreadsheet_llm.sheet_model import Sheet, load_xlsx
from python.spreadsheet_llm.structural_anchors import (
    extract_structural_anchors,
    find_row_anchors,
)

# ---------------------------------------------------------------------------
# Row sampling: within the Module-1-extracted sheet, detect contiguous table
# "bodies" (runs of data rows between header-like anchor rows) and, if a body
# run is long, keep only its head and tail verbatim.
# ---------------------------------------------------------------------------


@dataclass
class RowSamplingResult:
    kept_rows: List[int]  # rows (in the extracted sheet's coordinates) that survive
    omitted_runs: List[
        Tuple[int, int]
    ]  # (start, end) row ranges (extracted coords) that were dropped, for annotation


def sample_long_row_runs(
    sheet: Sheet,
    row_anchors_in_extracted_coords: set,
    sample_rows: int = 5,
    min_run_to_sample: int = 15,
) -> RowSamplingResult:
    """
    Walks the already-extracted (Module 1) sheet and finds maximal runs of
    consecutive rows that are NOT anchors (i.e. plain data-body rows,
    already thinned somewhat by extraction but can still be long if k is
    large or the sheet has few real boundaries). For any run longer than
    `min_run_to_sample`, keep only the first and last `sample_rows` rows of
    that run and drop the middle, recording the gap so it can be annotated
    in the output (e.g. "... 42 rows omitted ...").

    This is intentionally conservative: it never touches anchor rows
    (headers/boundaries), only long stretches of plain data rows.
    """
    n = sheet.n_rows
    kept = set()
    omitted_runs: List[Tuple[int, int]] = []

    run_start = None
    for r in range(1, n + 2):  # +1 sentinel to flush the last run
        is_anchor = r in row_anchors_in_extracted_coords or r > n
        if not is_anchor:
            if run_start is None:
                run_start = r
        else:
            if run_start is not None:
                run_end = r - 1
                run_len = run_end - run_start + 1
                if run_len > min_run_to_sample:
                    head = set(range(run_start, run_start + sample_rows))
                    tail = set(range(run_end - sample_rows + 1, run_end + 1))
                    kept |= head | tail
                    gap_start = run_start + sample_rows
                    gap_end = run_end - sample_rows
                    if gap_start <= gap_end:
                        omitted_runs.append((gap_start, gap_end))
                else:
                    kept |= set(range(run_start, run_end + 1))
                run_start = None
            if r <= n:
                kept.add(r)  # the anchor row itself

    kept_rows = sorted(kept)
    return RowSamplingResult(kept_rows=kept_rows, omitted_runs=omitted_runs)


def apply_row_sampling(sheet: Sheet, sampling: RowSamplingResult) -> Tuple[Sheet, dict]:
    """
    Builds a new Sheet containing only sampling.kept_rows (all columns kept),
    remapped to contiguous rows, and returns the row_map (old extracted-sheet
    row -> new sampled-sheet row) so omitted-run boundaries can be translated
    for annotation.
    """
    row_map = {orig: new for new, orig in enumerate(sampling.kept_rows, start=1)}
    new_grid = []
    for orig_r in sampling.kept_rows:
        new_row = []
        for c in range(1, sheet.n_cols + 1):
            src = sheet.get(orig_r, c)
            new_row.append(replace(src, row=row_map[orig_r]))
        new_grid.append(new_row)
    new_sheet = Sheet(new_grid, len(sampling.kept_rows), sheet.n_cols, name=sheet.name)
    return new_sheet, row_map


# ---------------------------------------------------------------------------
# Build "literal" regions (1x1 per non-empty cell -- no type-label
# aggregation) so Module 2's inverted-index step can consume them directly.
# Also inserts synthetic placeholder regions for omitted row-run gaps so the
# LLM can see that rows were skipped, and roughly how many.
# ---------------------------------------------------------------------------


def literal_regions(sheet: Sheet) -> List[AggregatedRegion]:
    regions = []
    for r in range(1, sheet.n_rows + 1):
        for c in range(1, sheet.n_cols + 1):
            cell = sheet.get(r, c)
            dtype = "Empty" if cell.is_empty else "String"
            regions.append(AggregatedRegion(r, c, r, c, dtype))
    return regions


def inject_gap_markers(
    sheet: Sheet, omitted_runs_new_coords: List[Tuple[int, int]]
) -> None:
    """
    Writes a human/LLM-readable placeholder string into column 1 of the row
    immediately before each gap, e.g. "... 37 similar rows omitted ...".
    Mutates the sheet's cell values in place. Assumes gaps have already been
    physically removed from the grid (i.e. this annotates at the seam, not
    inside a still-present range).
    """
    pass


# ---------------------------------------------------------------------------
# Formatting index: the inverted-index trick applied to *styles* instead of
# values. Cells are grouped by their visible formatting (fill color, font
# color, bold/italic/underline/strike, outlier font size), each group's cells
# are coalesced into maximal rectangles, and the result renders as compact
# (style|range) entries. Formatting often carries table semantics (header
# bands, color-coded status cells, bold totals) that pure values lose.
# ---------------------------------------------------------------------------


def _prevailing_font_size(sheet: Sheet) -> Optional[float]:
    """
    The modal font size across non-empty cells. Sizes equal to this are
    treated as the sheet's default and omitted from style signatures, so only
    outliers (titles, headers) get a "14pt" annotation.
    """
    counts = Counter(
        sheet.get(r, c).font_size
        for r in range(1, sheet.n_rows + 1)
        for c in range(1, sheet.n_cols + 1)
        if not sheet.get(r, c).is_empty and sheet.get(r, c).font_size
    )
    return counts.most_common(1)[0][0] if counts else None


def _coalesce_rectangles(
    coords: Set[Tuple[int, int]],
) -> List[Tuple[int, int, int, int]]:
    """
    Greedily merges a set of (row, col) coordinates into maximal rectangles:
    per-row horizontal runs first, then runs with an identical column span
    stacked across consecutive rows. Returns (top, left, bottom, right)
    tuples sorted by position.
    """
    cols_by_row: Dict[int, List[int]] = {}
    for r, c in coords:
        cols_by_row.setdefault(r, []).append(c)

    def _runs(cols: List[int]) -> List[Tuple[int, int]]:
        cols = sorted(cols)
        out = []
        start = prev = cols[0]
        for c in cols[1:]:
            if c == prev + 1:
                prev = c
            else:
                out.append((start, prev))
                start = prev = c
        out.append((start, prev))
        return out

    rects: List[Tuple[int, int, int, int]] = []
    open_rects: Dict[Tuple[int, int], Tuple[int, int]] = {}
    for r in sorted(cols_by_row):
        next_open: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for span in _runs(cols_by_row[r]):
            prev_rect = open_rects.get(span)
            if prev_rect is not None and prev_rect[1] == r - 1:
                next_open[span] = (prev_rect[0], r)
            else:
                next_open[span] = (r, r)
        for span, (top, last) in open_rects.items():
            if span not in next_open or next_open[span][0] != top:
                rects.append((top, span[0], last, span[1]))
        open_rects = next_open
    for span, (top, last) in open_rects.items():
        rects.append((top, span[0], last, span[1]))
    return sorted(rects)


def _rect_addr(top: int, left: int, bottom: int, right: int) -> str:
    a = f"{get_column_letter(left)}{top}"
    if (top, left) == (bottom, right):
        return a
    return f"{a}:{get_column_letter(right)}{bottom}"


def build_style_index(
    sheet: Sheet, base_font_size: Optional[float] = None
) -> Dict[str, List[str]]:
    """
    Builds {style_signature: [address_or_range, ...]} over every visibly
    styled cell, preserving first-seen order. Empty cells only count when
    they carry a fill (font styling on an empty cell is invisible); cells
    with fully default formatting are skipped, so a plain sheet yields {}.
    """
    groups: Dict[str, Set[Tuple[int, int]]] = {}
    order: List[str] = []
    for r in range(1, sheet.n_rows + 1):
        for c in range(1, sheet.n_cols + 1):
            cell = sheet.get(r, c)
            if cell.is_empty:
                sig = f"fill:{cell.fill_color}" if cell.fill_color else None
            else:
                sig = cell.style_signature(base_font_size)
            if not sig:
                continue
            if sig not in groups:
                groups[sig] = set()
                order.append(sig)
            groups[sig].add((r, c))
    return {
        sig: [_rect_addr(*rect) for rect in _coalesce_rectangles(groups[sig])]
        for sig in order
    }


# ---------------------------------------------------------------------------
# Full pipeline
# ---------------------------------------------------------------------------


@dataclass
class LLMContextResult:
    original_rows: int
    original_cols: int
    after_module1_rows: int
    after_module1_cols: int
    after_sampling_rows: int
    omitted_row_runs: List[Tuple[int, int]]
    compressed_text: str
    compressed_tokens: int
    vanilla_tokens: int
    compression_ratio: float


def _count_tokens(text: str) -> int:
    try:
        import tiktoken

        enc = tiktoken.get_encoding("cl100k_base")
        return len(enc.encode(text))
    except ImportError:
        return max(len(text.split()), len(text) // 4)


def _vanilla_token_estimate(sheet: Sheet) -> int:
    lines = []
    for r in range(1, sheet.n_rows + 1):
        parts = []
        for c in range(1, sheet.n_cols + 1):
            cell = sheet.get(r, c)
            val = "" if cell.is_empty else str(cell.value)
            parts.append(f"{cell.address},{val}")
        lines.append("|" + "|".join(parts) + "|")
    return _count_tokens("\n".join(lines))


def _find_module1_gaps(kept_rows: List[int]) -> List[Tuple[int, int]]:
    """
    Module 1 (structural-anchor extraction) can itself silently drop a long
    run of original rows. This scans the sorted list of ORIGINAL row numbers
    that Module 1 kept and returns the (start, end) original-row-number gaps
    between consecutive kept rows, so that loss can be annotated instead of
    silently disappearing.
    """
    gaps = []
    for prev, cur in zip(kept_rows, kept_rows[1:]):
        if cur - prev > 1:
            gaps.append((prev + 1, cur - 1))
    return gaps


def build_llm_context(
    path: str,
    sheet_name: Optional[str] = None,
    k: int = 4,
    sample_rows: int = 5,
    min_run_to_sample: int = 15,
    output_style: str = "paper",
    include_gap_notes: bool = True,
    enable_row_sampling: bool = True,
    enable_module1: bool = True,
) -> LLMContextResult:
    """
    Full pipeline, tuned for feeding an LLM that will generate a spreadsheet
    description. Real values are preserved.

    Flags:
      - include_gap_notes: whether to append [NOTE: rows ... omitted]
      - enable_row_sampling: whether to head/tail sample long body runs
      - enable_module1: whether to run structural-anchor extraction first
    """
    sheet = load_xlsx(path, sheet_name=sheet_name)
    vanilla_tokens = _vanilla_token_estimate(sheet)

    if enable_module1:
        extraction = extract_structural_anchors(sheet, k=k)
        extracted_sheet = extraction.sheet
        module1_gaps_original = _find_module1_gaps(extraction.kept_rows)
        inv_extract_row_map = {v: k_ for k_, v in extraction.row_map.items()}
    else:
        extracted_sheet = sheet
        module1_gaps_original = []
        inv_extract_row_map = {r: r for r in range(1, sheet.n_rows + 1)}

    anchors_in_extracted = find_row_anchors(extracted_sheet)

    if enable_row_sampling:
        sampling = sample_long_row_runs(
            extracted_sheet,
            anchors_in_extracted,
            sample_rows=sample_rows,
            min_run_to_sample=min_run_to_sample,
        )
        sampled_sheet, row_map = apply_row_sampling(extracted_sheet, sampling)
    else:
        sampling = RowSamplingResult(
            kept_rows=list(range(1, extracted_sheet.n_rows + 1)),
            omitted_runs=[],
        )
        sampled_sheet = extracted_sheet
        row_map = {r: r for r in range(1, extracted_sheet.n_rows + 1)}

    omitted_runs_original = list(module1_gaps_original)
    if enable_row_sampling:
        for start, end in sampling.omitted_runs:
            orig_start = inv_extract_row_map.get(start)
            orig_end = inv_extract_row_map.get(end)
            if orig_start is not None and orig_end is not None:
                omitted_runs_original.append((orig_start, orig_end))
    omitted_runs_original.sort()

    regions = literal_regions(sampled_sheet)
    index = build_inverted_index(sampled_sheet, regions)
    compressed_text = (
        render_paper_style(index)
        if output_style == "paper"
        else render_json_like(index)
    )

    if include_gap_notes and omitted_runs_original:
        gap_notes = "; ".join(
            f"rows {a}-{b} omitted ({b - a + 1} similar rows not shown)"
            for a, b in omitted_runs_original
        )
        compressed_text += f"\n\n[NOTE: {gap_notes}]"

    style_index = build_style_index(
        sampled_sheet, base_font_size=_prevailing_font_size(sampled_sheet)
    )
    if style_index:
        rendered_styles = (
            "".join(f"({sig}|{','.join(addrs)})" for sig, addrs in style_index.items())
            if output_style == "paper"
            else render_json_like(style_index)
        )
        compressed_text += f"\n\n[FORMATTING]\n{rendered_styles}"

    compressed_tokens = _count_tokens(compressed_text)
    ratio = vanilla_tokens / compressed_tokens if compressed_tokens else float("inf")

    return LLMContextResult(
        original_rows=sheet.n_rows,
        original_cols=sheet.n_cols,
        after_module1_rows=extracted_sheet.n_rows,
        after_module1_cols=extracted_sheet.n_cols,
        after_sampling_rows=sampled_sheet.n_rows,
        omitted_row_runs=omitted_runs_original,
        compressed_text=compressed_text,
        compressed_tokens=compressed_tokens,
        vanilla_tokens=vanilla_tokens,
        compression_ratio=ratio,
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Module 1 (structural anchors) + Module 2 (inverted index) "
        "spreadsheet encoder for LLM description-generation pipelines. "
        "Preserves real values (no data-format aggregation / type labels)."
    )
    parser.add_argument("xlsx_path", help="Path to the .xlsx file")
    parser.add_argument(
        "--sheet", default=None, help="Sheet name (default: active sheet)"
    )
    parser.add_argument(
        "--k",
        type=int,
        default=4,
        help="Structural anchor neighborhood radius (default: 4)",
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=5,
        help="How many head/tail rows to keep verbatim per long data run (default: 5)",
    )
    parser.add_argument(
        "--min-run-to-sample",
        type=int,
        default=15,
        help="A data-body run longer than this gets head/tail sampled (default: 15)",
    )
    parser.add_argument(
        "--format", choices=["paper", "json"], default="paper", help="Output text style"
    )
    parser.add_argument(
        "--no-gap-notes",
        action="store_true",
        help="Do not append [NOTE: rows ... omitted] annotations to the output",
    )
    parser.add_argument(
        "--no-row-sampling",
        action="store_true",
        help="Keep all rows after Module 1 extraction; do not head/tail sample long data runs",
    )
    parser.add_argument(
        "--no-module1",
        action="store_true",
        help="Skip structural-anchor extraction and keep all original rows",
    )

    args = parser.parse_args()

    result = build_llm_context(
        args.xlsx_path,
        sheet_name=args.sheet,
        k=args.k,
        sample_rows=args.sample_rows,
        min_run_to_sample=args.min_run_to_sample,
        output_style=args.format,
        include_gap_notes=not args.no_gap_notes,
        enable_row_sampling=not args.no_row_sampling,
        enable_module1=not args.no_module1,
    )

    # print("\n--- LLM-READY CONTEXT ---")
    print(result.compressed_text)


if __name__ == "__main__":
    main()
